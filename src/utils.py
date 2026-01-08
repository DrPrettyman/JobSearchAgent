import subprocess
import urllib.request
import re
import glob as globlib
from pathlib import Path


def resolve_paths(paths: list[str]) -> list[str]:
    """Resolve glob patterns in paths to actual file paths."""
    resolved_paths = []
    for path in paths:
        if "*" in path:
            matched = globlib.glob(path)
            resolved_paths.extend(p for p in matched if Path(p).is_file())
        else:
            resolved_paths.append(path)
    return list(set(resolved_paths))


def extract_text_from_docx(path: str) -> str:
    """Extract text from a Word document."""
    try:
        from docx import Document
        doc = Document(path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs)
    except ImportError:
        print(f"Warning: python-docx not installed. Cannot read {path}")
        return ""
    except Exception as e:
        print(f"Warning: Could not read Word doc {path}: {e}")
        return ""


def extract_text_from_pdf(path: str) -> str:
    """Extract text from a PDF file."""
    try:
        import pypdf
        reader = pypdf.PdfReader(path)
        text_parts = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                text_parts.append(text)
        return "\n\n".join(text_parts)
    except ImportError:
        print(f"Warning: pypdf not installed. Cannot read {path}")
        return ""
    except Exception as e:
        print(f"Warning: Could not read PDF {path}: {e}")
        return ""


def extract_text_from_tex(path: str) -> str:
    """Extract text from a LaTeX file, skipping the preamble."""
    try:
        with open(path, "r", encoding="utf-8") as f:
            content = f.read()

        # Find \begin{document} and extract content after it
        match = re.search(r'\\begin\{document\}', content)
        if match:
            content = content[match.end():]

        # Remove \end{document}
        content = re.sub(r'\\end\{document\}', '', content)

        # Remove common LaTeX commands but keep text
        # Remove comments
        content = re.sub(r'%.*$', '', content, flags=re.MULTILINE)
        # Remove \command{...} but keep content inside braces for text commands
        content = re.sub(r'\\(textbf|textit|emph|underline)\{([^}]*)\}', r'\2', content)
        # Remove \section{...} style commands but keep the title
        content = re.sub(r'\\(section|subsection|subsubsection|chapter|paragraph)\*?\{([^}]*)\}', r'\n\2\n', content)
        # Remove \item
        content = re.sub(r'\\item\s*', '• ', content)
        # Remove remaining simple commands
        content = re.sub(r'\\[a-zA-Z]+\*?(\[[^\]]*\])?(\{[^}]*\})?', '', content)
        # Remove environment markers
        content = re.sub(r'\\(begin|end)\{[^}]*\}', '', content)
        # Clean up braces
        content = re.sub(r'[{}]', '', content)
        # Clean up whitespace
        content = re.sub(r'\n{3,}', '\n\n', content)
        content = re.sub(r'[ \t]+', ' ', content)

        return content.strip()
    except Exception as e:
        print(f"Warning: Could not read LaTeX file {path}: {e}")
        return ""


def extract_text_from_csv(path: str) -> str:
    """Extract text from a CSV file by converting to JSON."""
    import csv
    import json
    try:
        with open(path, "r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        return json.dumps(rows, indent=2)
    except UnicodeDecodeError:
        try:
            with open(path, "r", encoding="latin-1", newline="") as f:
                reader = csv.DictReader(f)
                rows = list(reader)
            return json.dumps(rows, indent=2)
        except Exception as e:
            print(f"Warning: Could not read CSV {path}: {e}")
            return ""
    except Exception as e:
        print(f"Warning: Could not read CSV {path}: {e}")
        return ""


def extract_text_from_excel(path: str) -> str:
    """Extract text from an Excel file by converting to JSON."""
    import json
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        all_sheets = {}
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = []
            headers = None
            for row in sheet.iter_rows(values_only=True):
                # Skip empty rows
                if not any(cell is not None for cell in row):
                    continue
                if headers is None:
                    headers = [str(cell) if cell is not None else f"col_{j}" for j, cell in enumerate(row)]
                else:
                    row_dict = {headers[j]: cell for j, cell in enumerate(row) if j < len(headers)}
                    rows.append(row_dict)
            if rows:
                all_sheets[sheet_name] = rows
        return json.dumps(all_sheets, indent=2, default=str)
    except ImportError:
        print(f"Warning: openpyxl not installed. Cannot read {path}")
        return ""
    except Exception as e:
        print(f"Warning: Could not read Excel file {path}: {e}")
        return ""


def extract_text_from_json(path: str) -> str:
    """Extract text from a JSON file by pretty-printing it."""
    import json
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return json.dumps(data, indent=2)
    except Exception as e:
        print(f"Warning: Could not read JSON {path}: {e}")
        return ""


def extract_text_from_html(path: str) -> str:
    """Extract text from an HTML file."""
    try:
        with open(path, "r", encoding="utf-8") as f:
            html = f.read()
    except UnicodeDecodeError:
        try:
            with open(path, "r", encoding="latin-1") as f:
                html = f.read()
        except Exception as e:
            print(f"Warning: Could not read HTML {path}: {e}")
            return ""
    except Exception as e:
        print(f"Warning: Could not read HTML {path}: {e}")
        return ""

    # Remove script and style tags
    html_clean = re.sub(r'<script[^>]*>.*?</script>', '', html, flags=re.DOTALL)
    html_clean = re.sub(r'<style[^>]*>.*?</style>', '', html_clean, flags=re.DOTALL)
    # Remove HTML comments
    html_clean = re.sub(r'<!--.*?-->', '', html_clean, flags=re.DOTALL)
    # Remove HTML tags, replacing block elements with newlines
    html_clean = re.sub(r'</(p|div|br|h[1-6]|li|tr)>', '\n', html_clean, flags=re.IGNORECASE)
    html_clean = re.sub(r'<[^>]+>', ' ', html_clean)
    # Decode common HTML entities
    html_clean = html_clean.replace('&amp;', '&')
    html_clean = html_clean.replace('&lt;', '<')
    html_clean = html_clean.replace('&gt;', '>')
    html_clean = html_clean.replace('&quot;', '"')
    html_clean = html_clean.replace('&#39;', "'")
    html_clean = html_clean.replace('&nbsp;', ' ')
    # Clean up whitespace
    html_clean = re.sub(r'[ \t]+', ' ', html_clean)
    html_clean = re.sub(r'\n[ \t]+', '\n', html_clean)
    html_clean = re.sub(r'\n{3,}', '\n\n', html_clean)

    return html_clean.strip()


def extract_text_from_file(path: str) -> str:
    """Extract text from a file based on its extension."""
    path_obj = Path(path)
    ext = path_obj.suffix.lower()

    if ext == '.docx':
        text = extract_text_from_docx(path)
    elif ext == '.pdf':
        text = extract_text_from_pdf(path)
    elif ext == '.tex':
        text = extract_text_from_tex(path)
    elif ext == '.csv':
        text = extract_text_from_csv(path)
    elif ext in ('.xlsx', '.xls'):
        text = extract_text_from_excel(path)
    elif ext == '.json':
        text = extract_text_from_json(path)
    elif ext in ('.html', '.htm'):
        text = extract_text_from_html(path)
    else:
        # Plain text files (.txt, .md, etc.)
        try:
            with open(path, "r", encoding="utf-8") as f:
                text = f.read()
        except UnicodeDecodeError:
            try:
                with open(path, "r", encoding="latin-1") as f:
                    text = f.read()
            except Exception as e:
                print(f"Warning: Could not read {path}: {e}")
                text = ""
        except Exception as e:
            print(f"Warning: Could not read {path}: {e}")
            text = ""

    # Clean up excessive whitespace
    text = re.sub(r'[ \t]+', ' ', text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def combine_documents(paths: list[str]) -> list[dict]:
    """Combine multiple document files into a list of dicts.

    Returns:
        List of dicts, each with {"document_path": <path>, "content": <content>}
    """
    documents = []
    for path in resolve_paths(paths):
        content = extract_text_from_file(path)
        if content:
            documents.append({
                "document_path": path,
                "content": content
            })
    return documents


def combined_documents_as_string(documents: list[dict]) -> str:
    """Convert combined documents list to a single string for prompts."""
    if not documents:
        return ""
    parts = []
    for doc in documents:
        parts.append(f"=== {doc['document_path']} ===\n{doc['content']}")
    return "\n\n".join(parts)


def extract_url_slug(url: str) -> str:
    """Extract the last path segment from a URL.

    e.g.:
    - "https://github.com/torvalds" -> "torvalds"
    - "https://linkedin.com/in/prettyman/" -> "prettyman"
    - "username" -> "username"
    """
    return url.rstrip("/").split("/")[-1]


def run_claude(prompt: str, timeout: int = 120, tools: list[str] = None) -> tuple[bool, str]:
    """Run Claude CLI with a prompt. Returns (success, output).

    Args:
        prompt: The prompt to send to Claude
        timeout: Timeout in seconds
        tools: Optional list of tools to allow (e.g., ["WebSearch", "WebFetch"])
    """
    try:
        cmd = ["claude", "-p", prompt]
        if tools:
            cmd.extend(["--allowedTools", ",".join(tools)])
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=timeout
        )
        return result.returncode == 0, result.stdout
    except subprocess.TimeoutExpired:
        return False, "Timeout expired"
    except Exception as e:
        return False, str(e)
    
    
def scrape(url):
    """Fetch and extract text content from a job posting URL."""
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    }

    req = urllib.request.Request(url, headers=headers)

    try:
        response = urllib.request.urlopen(req, timeout=30)
        html = response.read().decode('utf-8')
    except Exception as e:
        raise RuntimeError(f"Error fetching URL: {e}")

    # Remove script and style tags
    html_clean = re.sub(r'<script[^>]*>.*?</script>', '', html, flags=re.DOTALL)
    html_clean = re.sub(r'<style[^>]*>.*?</style>', '', html_clean, flags=re.DOTALL)

    # Remove HTML tags, replacing with newlines
    html_clean = re.sub(r'<[^>]+>', '\n', html_clean)

    # Clean up whitespace
    html_clean = re.sub(r'\n+', '\n', html_clean)
    html_clean = re.sub(r'[ \t]+', ' ', html_clean)

    # Decode HTML entities
    html_clean = html_clean.replace('&amp;', '&')
    html_clean = html_clean.replace('&lt;', '<')
    html_clean = html_clean.replace('&gt;', '>')
    html_clean = html_clean.replace('&quot;', '"')
    html_clean = html_clean.replace('&#39;', "'")
    html_clean = html_clean.replace('&nbsp;', ' ')

    # Extract lines and clean
    lines = [l.strip() for l in html_clean.split('\n') if l.strip()]

    return '\n'.join(lines)


def summarize_source_documents(combined_docs: str) -> str:
    """Generate a short summary of source documents using Claude.

    Args:
        combined_docs: Combined text from all source documents

    Returns:
        Short summary string, or empty string on failure
    """
    if not combined_docs or len(combined_docs) < 50:
        return ""

    # Truncate to avoid token limits
    docs_truncated = combined_docs[:12000]

    prompt = f"""Summarize this person's professional background in 2-3 sentences.
Focus on: key skills, experience level, and main areas of expertise.
Be concise and factual.

Documents:
{docs_truncated}"""

    success, response = run_claude(prompt, timeout=60)
    if not success:
        return ""

    return response.strip()


def summarize_online_presence(online_presence: list[dict]) -> str:
    """Generate a short summary of online presence using Claude.

    Args:
        online_presence: List of dicts with site, time_fetched, content

    Returns:
        Short summary string, or empty string on failure
    """
    if not online_presence:
        return ""

    # Combine all content
    combined = "\n\n".join([
        f"--- {entry.get('site', 'Unknown')} ---\n{entry.get('content', '')}"
        for entry in online_presence
    ])

    if len(combined) < 50:
        return ""

    # Truncate to avoid token limits
    combined_truncated = combined[:12000]

    prompt = f"""Summarize this person's online presence in 2-3 sentences.
Focus on: notable projects, public contributions, and professional highlights.
Be concise and factual.

Online profiles:
{combined_truncated}"""

    success, response = run_claude(prompt, timeout=60)
    if not success:
        return ""

    return response.strip()
